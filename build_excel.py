#!/usr/bin/env python3
"""
Costruisce espositori_tuttofood_2026.xlsx a partire da espositori_tuttofood_2026.csv.

Foglio "Espositori":
  - tutte le righe del CSV
  - header in grassetto, freeze pane sulla riga 1, autofilter
  - larghezza colonne adattata al contenuto (con cap)

Foglio "Statistiche":
  - totale espositori
  - top 20 città
  - top 20 paesi
  - distribuzione per regione italiana
"""
import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent
CSV_PATH = OUT_DIR / "espositori_tuttofood_2026.csv"
XLSX_PATH = OUT_DIR / "espositori_tuttofood_2026.xlsx"

MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 10
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12)


def read_rows():
    with CSV_PATH.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []
    return fieldnames, rows


def autosize(ws, headers, rows):
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for r in rows:
            v = r.get(h, "") or ""
            if len(v) > max_len:
                max_len = len(v)
        width = min(MAX_COL_WIDTH, max(MIN_COL_WIDTH, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_espositori(wb, headers, rows):
    ws = wb.active
    ws.title = "Espositori"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # freeze + autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    autosize(ws, headers, rows)
    ws.row_dimensions[1].height = 22


def add_section(ws, title):
    ws.append([])
    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = SECTION_FONT


def add_table(ws, header_row, data_rows):
    ws.append(header_row)
    for c in ws[ws.max_row]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for r in data_rows:
        ws.append(r)


def build_statistiche(wb, rows):
    ws = wb.create_sheet("Statistiche")

    # 1. Totale
    ws.append(["Totale espositori", len(rows)])
    ws.cell(row=1, column=1).font = SECTION_FONT
    coesp = sum(1 for r in rows if (r.get("coespositore") or "").strip() == "1")
    ws.append(["di cui co-espositori", coesp])
    ws.append(["Espositori principali", len(rows) - coesp])
    italiani = sum(1 for r in rows if (r.get("regione") or "").strip())
    ws.append(["Espositori italiani", italiani])
    ws.append(["Espositori esteri", len(rows) - italiani])

    # 2. Top 20 città
    add_section(ws, "Top 20 città")
    citta_counter = Counter(
        (r.get("citta") or "").strip().upper() for r in rows if (r.get("citta") or "").strip()
    )
    add_table(ws, ["Città", "Espositori"], citta_counter.most_common(20))

    # 3. Top 20 paesi
    add_section(ws, "Top 20 paesi")
    paese_counter = Counter(
        (r.get("paese") or "").strip().upper() for r in rows if (r.get("paese") or "").strip()
    )
    add_table(ws, ["Paese", "Espositori"], paese_counter.most_common(20))

    # 4. Regioni italiane
    add_section(ws, "Distribuzione per regione italiana")
    reg_counter = Counter(
        (r.get("regione") or "").strip() for r in rows if (r.get("regione") or "").strip()
    )
    add_table(ws, ["Regione", "Espositori"], sorted(reg_counter.items(), key=lambda x: -x[1]))

    # 5. Padiglioni
    add_section(ws, "Distribuzione per padiglione (Tuttofood Milano)")
    pad_counter = Counter(
        (r.get("padiglione") or "").strip() for r in rows if (r.get("padiglione") or "").strip()
    )
    add_table(
        ws,
        ["Padiglione", "Espositori"],
        sorted(pad_counter.items(), key=lambda x: (len(x[0]), x[0])),
    )

    # 6. Coverage campi (% non-vuoti)
    add_section(ws, "Coverage campi (% righe valorizzate)")
    add_table(ws, ["Campo", "% valorizzato", "N. valorizzati"], [])
    n = max(1, len(rows))
    for h in (rows[0].keys() if rows else []):
        filled = sum(1 for r in rows if (r.get(h) or "").strip())
        ws.append([h, f"{filled / n * 100:.1f}%", filled])

    # widths
    for col_idx in range(1, 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = 35


def main():
    if not CSV_PATH.exists():
        raise SystemExit(f"CSV non trovato: {CSV_PATH}")
    headers, rows = read_rows()
    print(f"[xlsx] {len(rows)} righe da {CSV_PATH.name}")
    wb = Workbook()
    build_espositori(wb, headers, rows)
    build_statistiche(wb, rows)
    wb.save(XLSX_PATH)
    print(f"[xlsx] Salvato: {XLSX_PATH}")


if __name__ == "__main__":
    main()
